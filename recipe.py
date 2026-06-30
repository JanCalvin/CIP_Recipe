import streamlit as st
import openpyxl
from io import BytesIO
import os
st.title("Upload & Transform Excel CIP Matrix")
st.subheader("Semangat Ges! 🚀")

uploaded_file = st.file_uploader("Upload file Excel (WKWK.xlsx)", type=["xlsx"])

HEADER_ROW = 5  # baris berisi TDS Code (SKU_After) di kolom F dst
SKU_COL = 5     # kolom E berisi TDS Code (SKU) per baris


def transform_workbook(file):
    wb_src = openpyxl.load_workbook(file, data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    summary = {}

    for sheet_name in wb_src.sheetnames:
        ws = wb_src[sheet_name]

        # Kumpulkan SKU_After dari header_row, kolom F dst
        col_skus = {}
        for col in range(SKU_COL + 1, ws.max_column + 1):
            val = ws.cell(row=HEADER_ROW, column=col).value
            if val is not None and str(val).strip() != "":
                col_skus[col] = str(val).strip()

        out_rows = [["SKU", "SKU_After", "Recipe"]]

        for row in range(HEADER_ROW + 1, ws.max_row + 1):
            row_sku = ws.cell(row=row, column=SKU_COL).value
            if row_sku is None or str(row_sku).strip() == "":
                continue
            row_sku = str(row_sku).strip()

            for col, col_sku in col_skus.items():
                recipe = ws.cell(row=row, column=col).value
                if recipe is None or str(recipe).strip() == "":
                    continue
                recipe = str(recipe).strip()
                if recipe.lower() == "no cip":
                    continue
                out_rows.append([row_sku, col_sku, recipe])

        ws_out = wb_out.create_sheet(title=sheet_name)
        for r in out_rows:
            ws_out.append(r)

        # bold header
        for cell in ws_out[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # lebar kolom
        for col_letter in ["A", "B", "C"]:
            ws_out.column_dimensions[col_letter].width = 18

        summary[sheet_name] = len(out_rows) - 1  # jumlah data row (tanpa header)

    return wb_out, summary

if uploaded_file:
    st.success("File berhasil diupload!")

    try:
        wb_hasil, summary = transform_workbook(uploaded_file)
    except Exception as e:
        st.error(f"Gagal memproses file: {e}")
        st.stop()

    # Ambil nama sheet pertama buat nama file output
    first_sheet = wb_hasil.sheetnames[0]
    output_filename = f"{first_sheet}.xlsx"

    st.write("Ringkasan hasil per sheet:")
    for sheet_name, n_rows in summary.items():
        st.write(f"- **{sheet_name}**: {n_rows} baris")

    # Preview sheet pertama
    ws_preview = wb_hasil[first_sheet]
    preview_rows = list(ws_preview.iter_rows(min_row=1, max_row=6, values_only=True))
    st.write(f"Preview sheet '{first_sheet}':")
    st.table(preview_rows)

    output = BytesIO()
    wb_hasil.save(output)
    output.seek(0)

    st.download_button(
        label="Download hasil Excel",
        data=output,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )